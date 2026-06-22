import os
import uuid
import bcrypt
import smtplib
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.db import IntegrityError
from django.urls import reverse
from .models import Usuario
from django.db.models import Q
from datetime import date

logger = logging.getLogger(__name__)

def login_view(request):
    if request.user.is_authenticated:
        return redirect('panel_admin')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('clave', '')
        if not email or not password:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'login.html')
        try:
            user = authenticate(request, email=email, password=password)
        except Exception as e:
            logger.exception("Error al autenticar usuario %s: %s", email, e)
            messages.error(request, 'Error al iniciar sesión. Intenta de nuevo.')
            return render(request, 'login.html')
        if user is not None:
            try:
                login(request, user)
            except Exception as e:
                logger.exception("Error al iniciar sesión para %s: %s", email, e)
                messages.error(request, 'Error al iniciar sesión. Intenta de nuevo.')
                return render(request, 'login.html')
            return redirect('panel_admin')
        else:
            messages.error(request, 'Credenciales inválidas o cuenta no activada.')

    return render(request, 'login.html')

def logout_view(request):
    try:
        logout(request)
    except Exception as e:
        logger.exception("Error al cerrar sesión: %s", e)
    return redirect('login')

@login_required
def panel_admin(request):
    usuarios = Usuario.objects.all()

    query = request.GET.get('q', '').strip()
    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) | Q(email__icontains=query)
        )

    total_filtrados = usuarios.count()
    total_usuarios = Usuario.objects.count()
    rrhh_activos = Usuario.objects.filter(rol='ROLE_RRHH', activo=True).count()
    candidatos = Usuario.objects.filter(rol='ROLE_CANDIDATO').count()
    pendientes = Usuario.objects.filter(rol='ROLE_CANDIDATO', activo=False).count()

    context = {
        'usuarios': usuarios,
        'total_usuarios': total_usuarios,
        'total_filtrados': total_filtrados,
        'rrhh_activos': rrhh_activos,
        'candidatos': candidatos,
        'pendientes': pendientes,
    }
    return render(request, 'admin.html', context)

@login_required
def crear_rrhh(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        username = request.POST.get('username', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        clave = request.POST.get('clave', '')

        if not email or not username or not apellido or not clave:
            return HttpResponseRedirect(reverse('panel_admin') + '?error=campos_vacios')
        if len(clave) < 8:
            return HttpResponseRedirect(reverse('panel_admin') + '?error=clave_corta')

        if Usuario.objects.filter(email=email).exists():
            return redirect('/admin/?error=duplicado')

        token = str(uuid.uuid4())
        try:
            usuario = Usuario(
                username=username,
                apellido=apellido,
                email=email,
                clave=bcrypt.hashpw(clave.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
                rol='ROLE_RRHH',
                activo=False,
                tokenactivacion=token,
            )
            usuario.save()
        except Exception as e:
            logger.exception("Error al crear RRHH %s: %s", email, e)
            return HttpResponseRedirect(reverse('panel_admin') + '?error=duplicado')

        try:
            enlace = request.build_absolute_uri(f'/activar/?token={token}')
        except Exception as e:
            logger.exception("Error al construir URI para %s: %s", email, e)
            enlace = f'http://localhost:8000/activar/?token={token}'

        asunto = 'Activa tu cuenta en FLOWMATIC'
        mensaje = f'''
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:32px;background:#f9f9f9;border-radius:12px;">
            <div style="text-align:center;margin-bottom:28px;">
                <h1 style="color:#0D9488;font-size:24px;margin:0;">FLOWMATIC</h1>
                <p style="color:#666;font-size:14px;">Gestión de Reclutamiento</p>
            </div>
            <div style="background:#fff;padding:32px;border-radius:8px;">
                <h2 style="color:#1a1a2e;font-size:18px;margin:0 0 12px;">Hola, {username}!</h2>
                <p style="color:#555;font-size:14px;line-height:1.6;">Te han registrado como personal RRHH en FLOWMATIC. Para activar tu cuenta, haz clic en el botón:</p>
                <div style="text-align:center;margin:28px 0;">
                    <a href="{enlace}" style="background:#0D9488;color:#fff;padding:14px 36px;border-radius:8px;text-decoration:none;font-size:15px;font-weight:600;display:inline-block;">Activar mi cuenta</a>
                </div>
                <p style="color:#999;font-size:12px;">Si no esperabas este correo, ignóralo.</p>
            </div>
        </div>
        '''
        try:
            send_mail(asunto, '', settings.DEFAULT_FROM_EMAIL, [email], html_message=mensaje, fail_silently=False)
        except smtplib.SMTPException:
            return HttpResponseRedirect(reverse('panel_admin') + '?envio_exitoso=1&email_fallo=1')

        return HttpResponseRedirect(reverse('panel_admin') + '?envio_exitoso=1')

    return redirect('panel_admin')


@login_required
def eliminar_usuario(request, usuario_id):
    try:
        usuario = get_object_or_404(Usuario, id=usuario_id)
        usuario.delete()
    except Exception as e:
        logger.exception("Error al eliminar usuario %s: %s", usuario_id, e)
        return redirect('panel_admin')
    return redirect('panel_admin')


@login_required
def editar_usuario(request):
    if request.method == 'POST':
        usuario_id = request.POST.get('id')
        if not usuario_id:
            return redirect('panel_admin')

        usuario = get_object_or_404(Usuario, id=usuario_id)

        username = request.POST.get('username', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        email = request.POST.get('email', '').strip()

        if not username or not apellido or not email:
            return HttpResponseRedirect(reverse('panel_admin') + '?error=campos_vacios')

        usuario.username = username
        usuario.apellido = apellido
        usuario.email = email
        usuario.telefono = request.POST.get('telefono', '').strip() or None

        nueva_clave = request.POST.get('nuevaClave', '')
        if nueva_clave and len(nueva_clave) >= 8:
            try:
                usuario.clave = bcrypt.hashpw(
                    nueva_clave.encode('utf-8'),
                    bcrypt.gensalt()
                ).decode('utf-8')
            except Exception as e:
                logger.exception("Error al hashear clave para usuario %s: %s", email, e)
                return HttpResponseRedirect(reverse('panel_admin') + '?error=hash_fallo')

        try:
            usuario.save()
        except IntegrityError:
            return HttpResponseRedirect(reverse('panel_admin') + '?error=duplicado')
        return HttpResponseRedirect(reverse('panel_admin') + '?editado=1')

    return redirect('panel_admin')

@login_required
def exportar_excel(request):
    try:
        cols_seleccionadas = request.GET.getlist('col')
        if not cols_seleccionadas:
            cols_seleccionadas = ['id', 'username', 'apellido', 'email', 'rol', 'estado']
            
        query = request.GET.get('q', '')
        rol = request.GET.get('rol', '')
        usuarios = Usuario.objects.all()
        if query:
            usuarios = usuarios.filter(Q(username__icontains=query) | Q(email__icontains=query))
        if rol:
            usuarios = usuarios.filter(rol=rol)
        mapeo = {
            'id': {'label': 'ID', 'attr': 'id'},
            'username': {'label': 'Username', 'attr': 'username'},
            'apellido': {'label': 'Apellido', 'attr': 'apellido'},
            'email': {'label': 'Email', 'attr': 'email'},
            'rol': {'label': 'Rol', 'attr': 'rol'},
            'estado': {'label': 'Estado', 'attr': None}
        }

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Usuarios'
        ultima_col = len(cols_seleccionadas)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center')

        # Row 1 — Logo
        try:
            ruta_logo = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png')
            if os.path.exists(ruta_logo):
                img = Image(ruta_logo)
                sheet.row_dimensions[1].height = 55
                img.height = 55
                img.width = 195
                sheet.add_image(img, 'A1')
        except Exception as e:
            logger.exception("Error al agregar logo al Excel: %s", e)

        # Row 2 — Title
        if ultima_col > 1:
            try:
                sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_col)
            except Exception as e:
                logger.exception("Error al mergear celdas del título: %s", e)
        title_cell = sheet.cell(row=2, column=1, value='FLOWMATIC · Gestión de Talento')
        title_cell.font = Font(name='Calibri', size=16, bold=True, color='0D9488')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[2].height = 30

        # Row 3 — Date / filter info
        if ultima_col > 1:
            try:
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ultima_col)
            except Exception as e:
                logger.exception("Error al mergear celdas de fecha: %s", e)
        info = f'Generado: {date.today().strftime("%d/%m/%Y")}'
        if query:
            info += f' | Búsqueda: "{query}"'
        info_cell = sheet.cell(row=3, column=1, value=info)
        info_cell.font = Font(name='Calibri', size=10, color='6B7280', italic=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[3].height = 18

        # Row 4 — Headers
        header_row = 4
        header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
        header_font = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
        for col_idx, col_key in enumerate(cols_seleccionadas, 1):
            cell = sheet.cell(row=header_row, column=col_idx, value=mapeo[col_key]['label'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        sheet.row_dimensions[header_row].height = 24

        # Row 5+ — Data with zebra striping and status colors
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        active_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        active_font = Font(name='Calibri', color='065F46', bold=True, size=11)
        pending_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        pending_font = Font(name='Calibri', color='92400E', bold=True, size=11)
        data_font = Font(name='Calibri', size=11)
        role_map = {
            'ROLE_ADMINISTRADOR': 'Administrador',
            'ROLE_RRHH': 'RRHH',
            'ROLE_CANDIDATO': 'Candidato',
        }

        for row_offset, u in enumerate(usuarios):
            row_idx = header_row + 1 + row_offset
            for col_idx, col_key in enumerate(cols_seleccionadas, 1):
                if col_key == 'estado':
                    val = 'Activo' if u.activo else 'Pendiente'
                elif col_key == 'rol':
                    val = role_map.get(u.rol, str(u.rol))
                else:
                    val = getattr(u, mapeo[col_key]['attr'])

                cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center
                cell.border = thin_border

                if col_key == 'estado':
                    if val == 'Activo':
                        cell.fill = active_fill
                        cell.font = active_font
                    else:
                        cell.fill = pending_fill
                        cell.font = pending_font
                else:
                    if row_offset % 2 == 1:
                        cell.fill = zebra_fill
                    cell.font = data_font

        # — Auto‑ajuste de ancho de columnas —
        for col_cells in sheet.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for c in col_cells:
                try:
                    if c.value and len(str(c.value)) > max_length:
                        max_length = len(str(c.value))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_length + 3, 40)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="usuarios_reporte.xlsx"'
        workbook.save(response)
        return response
    except Exception as e:
        logger.exception("Error al exportar Excel: %s", e)
        return redirect('panel_admin')

def registro_candidato(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        username = request.POST.get('username', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        clave = request.POST.get('clave', '')
        telefono = request.POST.get('telefono', '').strip() or None

        if not email or not username or not apellido or not clave:
            return render(request, 'registro-candidato.html', {
                'error_campos': True,
            })
        if len(clave) < 8:
            return render(request, 'registro-candidato.html', {
                'error_clave_corta': True,
            })

        try:
            if Usuario.objects.filter(email=email).exists():
                return render(request, 'registro-candidato.html', {
                    'error_duplicado': True,
                })
        except Exception as e:
            logger.exception("Error al verificar email duplicado %s: %s", email, e)
            return render(request, 'registro-candidato.html', {
                'error_duplicado': True,
            })

        token = str(uuid.uuid4())
        try:
            usuario = Usuario(
                username=username,
                apellido=apellido,
                email=email,
                clave=bcrypt.hashpw(clave.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
                telefono=telefono,
                rol='ROLE_CANDIDATO',
                activo=False,
                tokenactivacion=token,
            )
            usuario.save()
        except Exception as e:
            logger.exception("Error al registrar candidato %s: %s", email, e)
            return render(request, 'registro-candidato.html', {
                'error_duplicado': True,
            })

        try:
            enlace = request.build_absolute_uri(f'/activar/?token={token}')
        except Exception as e:
            logger.exception("Error al construir URI para %s: %s", email, e)
            enlace = f'http://localhost:8000/activar/?token={token}'

        asunto = 'Activa tu cuenta en FLOWMATIC'
        mensaje = f'''
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:32px;background:#f9f9f9;border-radius:12px;">
            <div style="text-align:center;margin-bottom:28px;">
                <h1 style="color:#0D9488;font-size:24px;margin:0;">FLOWMATIC</h1>
                <p style="color:#666;font-size:14px;">Gestión de Reclutamiento</p>
            </div>
            <div style="background:#fff;padding:32px;border-radius:8px;">
                <h2 style="color:#1a1a2e;font-size:18px;margin:0 0 12px;">Hola, {username}!</h2>
                <p style="color:#555;font-size:14px;line-height:1.6;">Gracias por registrarte. Para activar tu cuenta y empezar a usar FLOWMATIC, haz clic en el botón:</p>
                <div style="text-align:center;margin:28px 0;">
                    <a href="{enlace}" style="background:#0D9488;color:#fff;padding:14px 36px;border-radius:8px;text-decoration:none;font-size:15px;font-weight:600;display:inline-block;">Activar mi cuenta</a>
                </div>
                <p style="color:#999;font-size:12px;">Si no creaste esta cuenta, ignora este mensaje.</p>
            </div>
        </div>
        '''
        try:
            send_mail(asunto, '', settings.DEFAULT_FROM_EMAIL, [email], html_message=mensaje, fail_silently=False)
        except smtplib.SMTPException:
            pass

        return redirect(f'{reverse("registro_candidato")}?pendiente=1')

    context = {
        'mensaje_pendiente': request.GET.get('pendiente') == '1',
    }
    return render(request, 'registro-candidato.html', context)


def activar_cuenta(request):
    token = request.GET.get('token')
    if not token:
        return render(request, 'activacion.html', {'token_invalido': True})

    try:
        usuario = Usuario.objects.get(tokenactivacion=token)
        usuario.activo = True
        usuario.tokenactivacion = None
        usuario.save()
        return render(request, 'activacion.html', {'activacion_exitosa': True})
    except Usuario.DoesNotExist:
        return render(request, 'activacion.html', {'token_invalido': True})


def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            return render(request, 'forgot-password.html', {'error_campos': True})

        try:
            usuario = Usuario.objects.get(email=email)
        except Usuario.DoesNotExist:
            return redirect('/forgot-password/?success=1')

        token = str(uuid.uuid4())
        usuario.tokenactivacion = token
        if not usuario.apellido or not usuario.apellido.strip():
            usuario.apellido = 'N/A'
        try:
            usuario.save()
        except Exception as e:
            logger.exception("Error al guardar token de recuperación para %s: %s", email, e)
            return redirect('/forgot-password/?success=1')

        try:
            enlace = request.build_absolute_uri(f'/reset-password?token={token}')
        except Exception as e:
            logger.exception("Error al construir URI para %s: %s", usuario.email, e)
            enlace = f'http://localhost:8000/reset-password?token={token}'

        asunto = 'Restablece tu contraseña en FLOWMATIC'
        mensaje = f'''
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:32px;background:#f9f9f9;border-radius:12px;">
            <div style="text-align:center;margin-bottom:28px;">
                <h1 style="color:#0D9488;font-size:24px;margin:0;">FLOWMATIC</h1>
                <p style="color:#666;font-size:14px;">Gestión de Reclutamiento</p>
            </div>
            <div style="background:#fff;padding:32px;border-radius:8px;">
                <h2 style="color:#1a1a2e;font-size:18px;margin:0 0 12px;">Hola, {usuario.username}!</h2>
                <p style="color:#555;font-size:14px;line-height:1.6;">Recibimos una solicitud para restablecer la contraseña de tu cuenta en FLOWMATIC.</p>
                <p style="color:#555;font-size:14px;line-height:1.6;">Haz clic en el siguiente botón para crear una nueva contraseña:</p>
                <div style="text-align:center;margin:28px 0;">
                    <a href="{enlace}" style="background:#0D9488;color:#fff;padding:14px 36px;border-radius:8px;text-decoration:none;font-size:15px;font-weight:600;display:inline-block;">Restablecer contraseña</a>
                </div>
                <p style="color:#999;font-size:12px;">Si no solicitaste el restablecimiento de contraseña, ignora este mensaje.</p>
            </div>
        </div>
        '''
        try:
            send_mail(asunto, '', settings.DEFAULT_FROM_EMAIL, [email], html_message=mensaje, fail_silently=False)
        except smtplib.SMTPException:
            pass

        return redirect('/forgot-password/?success=1')

    success = request.GET.get('success') == '1'
    return render(request, 'forgot-password.html', {'success': success})


def reset_password(request):
    success = request.GET.get('success')

    if success:
        return render(request, 'reset-password.html', {'success': True})

    if request.method == 'POST':
        token = request.POST.get('token', '').strip()
        password = request.POST.get('password', '')

        if not token:
            return render(request, 'reset-password.html', {'error_token': True})

        if len(password) < 8:
            return render(request, 'reset-password.html', {
                'token': token,
                'error_password': True,
            })

        try:
            usuario = Usuario.objects.get(tokenactivacion=token)
        except Usuario.DoesNotExist:
            return redirect('/forgot-password/?error_token=1')

        try:
            usuario.clave = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        except Exception as e:
            logger.exception("Error al hashear nueva clave: %s", e)
            return redirect('/forgot-password/?error_token=1')

        usuario.tokenactivacion = None
        try:
            usuario.save()
        except Exception as e:
            logger.exception("Error al guardar nueva clave para token: %s", e)
            return redirect('/forgot-password/?error_token=1')

        return redirect('/reset-password?success=1')

    token = request.GET.get('token', '').strip()
    if not token:
        return render(request, 'reset-password.html', {'error_token': True})

    return render(request, 'reset-password.html', {'token': token})
